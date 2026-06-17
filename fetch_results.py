#!/usr/bin/env python3
"""
fetch_results.py — Automatic FIFA WC 2026 result fetcher.

Data sources (tried in order):
  1. football-data.org (primary, works everywhere including GitHub Actions)
     Requires free API key → https://www.football-data.org/client/register
     Set env var: FOOTBALL_DATA_API_KEY=your_key
  2. worldcup26.ir (fallback, works locally only — blocked by Azure/GitHub IPs)

Usage:
    python3 fetch_results.py                  # fetch + update Excel + push
    python3 fetch_results.py --dry-run        # fetch only, show what would change
    python3 fetch_results.py --no-push        # update Excel but skip Supabase push
    python3 fetch_results.py --ticker-only    # only update the ticker text in Supabase
"""

import os, sys, json, urllib.request, urllib.error, argparse, datetime, subprocess, time

# ─── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DESKTOP = os.path.expanduser("~/Desktop/FIFAWC2026")
BASE = BASE_DESKTOP if os.path.exists(BASE_DESKTOP) else SCRIPT_DIR
MODEL = os.path.join(BASE, "FIFAWC2026_Model.xlsx")

# ─── API endpoints ────────────────────────────────────────────────────────────
FOOTBALLDATA_URL = "https://api.football-data.org/v4/competitions/WC/matches"
WORLDCUP26_URL   = "https://worldcup26.ir/get/games"

# ─── Team name mappings (API name → Excel UPPER name) ─────────────────────────
# football-data.org names
FOOTBALLDATA_TEAM_MAP = {
    "Mexico":                           "MEXICO",
    "South Africa":                     "SOUTH AFRICA",
    "Korea Republic":                   "SOUTH KOREA",
    "Czechia":                          "CZECHIA",
    "Canada":                           "CANADA",
    "Bosnia and Herzegovina":           "BOSNIA H.",
    "Qatar":                            "QATAR",
    "Switzerland":                      "SWITZERLAND",
    "Brazil":                           "BRAZIL",
    "Morocco":                          "MOROCCO",
    "Haiti":                            "HAITI",
    "Scotland":                         "SCOTLAND",
    "USA":                              "USA",
    "United States":                    "USA",
    "Paraguay":                         "PARAGUAY",
    "Australia":                        "AUSTRALIA",
    "Turkey":                           "TURKEY",
    "Türkiye":                          "TURKEY",
    "Germany":                          "GERMANY",
    "Curaçao":                          "CURAÇAO",
    "Netherlands":                      "NETHERLANDS",
    "Japan":                            "JAPAN",
    "Sweden":                           "SWEDEN",
    "Tunisia":                          "TUNISIA",
    "Iran":                             "IRAN",
    "New Zealand":                      "NEW ZEALAND",
    "Belgium":                          "BELGIUM",
    "Egypt":                            "EGYPT",
    "Spain":                            "SPAIN",
    "Cape Verde":                       "CAPO VERDE",
    "Saudi Arabia":                     "SAUDI ARABIA",
    "Uruguay":                          "URUGUAY",
    "France":                           "FRANCE",
    "Senegal":                          "SENEGAL",
    "Iraq":                             "IRAQ",
    "Norway":                           "NORWAY",
    "Argentina":                        "ARGENTINA",
    "Algeria":                          "ALGERIA",
    "Austria":                          "AUSTRIA",
    "Jordan":                           "JORDAN",
    "Ivory Coast":                      "IVORY COAST",
    "Côte d'Ivoire":                    "IVORY COAST",
    "Ecuador":                          "ECUADOR",
    "Portugal":                         "PORTUGAL",
    "DR Congo":                         "DR CONGO",
    "Democratic Republic of Congo":     "DR CONGO",
    "Uzbekistan":                       "UZBEKISTAN",
    "Colombia":                         "COLOMBIA",
    "England":                          "ENGLAND",
    "Croatia":                          "CROATIA",
    "Ghana":                            "GHANA",
    "Panama":                           "PANAMA",
}

# worldcup26.ir names (same keys, slightly different spellings)
WORLDCUP26_TEAM_MAP = {
    **FOOTBALLDATA_TEAM_MAP,
    "Czech Republic":                   "CZECHIA",
    "Democratic Republic of the Congo": "DR CONGO",
    "South Korea":                      "SOUTH KOREA",
    "United States":                    "USA",
}

# ─── Excel group layout ────────────────────────────────────────────────────────
GH = {g: r for g, r in zip("ABCDEFGHIJKL", [4, 11, 18, 25, 32, 39, 46, 53, 60, 67, 74, 81])}

# ─── Display names for ticker (Excel name → readable ticker name) ─────────────
DISPLAY_NAMES = {
    "MEXICO":        "Mexico",
    "SOUTH AFRICA":  "South Africa",
    "SOUTH KOREA":   "South Korea",
    "CZECHIA":       "Czechia",
    "CANADA":        "Canada",
    "BOSNIA H.":     "Bosnia Herzegovina",
    "QATAR":         "Qatar",
    "SWITZERLAND":   "Switzerland",
    "BRAZIL":        "Brazil",
    "MOROCCO":       "Morocco",
    "HAITI":         "Haiti",
    "SCOTLAND":      "Scotland",
    "USA":           "USA",
    "PARAGUAY":      "Paraguay",
    "AUSTRALIA":     "Australia",
    "TURKEY":        "Turkey",
    "GERMANY":       "Germany",
    "CURAÇAO":       "Curaçao",
    "NETHERLANDS":   "Netherlands",
    "JAPAN":         "Japan",
    "SWEDEN":        "Sweden",
    "TUNISIA":       "Tunisia",
    "IRAN":          "Iran",
    "NEW ZEALAND":   "New Zealand",
    "BELGIUM":       "Belgium",
    "EGYPT":         "Egypt",
    "SPAIN":         "Spain",
    "CAPO VERDE":    "Cape Verde",
    "SAUDI ARABIA":  "Saudi Arabia",
    "URUGUAY":       "Uruguay",
    "FRANCE":        "France",
    "SENEGAL":       "Senegal",
    "IRAQ":          "Iraq",
    "NORWAY":        "Norway",
    "ARGENTINA":     "Argentina",
    "ALGERIA":       "Algeria",
    "AUSTRIA":       "Austria",
    "JORDAN":        "Jordan",
    "IVORY COAST":   "Ivory Coast",
    "ECUADOR":       "Ecuador",
    "PORTUGAL":      "Portugal",
    "DR CONGO":      "DR Congo",
    "UZBEKISTAN":    "Uzbekistan",
    "COLOMBIA":      "Colombia",
    "ENGLAND":       "England",
    "CROATIA":       "Croatia",
    "GHANA":         "Ghana",
    "PANAMA":        "Panama",
}

# 3-letter codes for next game HUD display
SHORT_CODES = {
    "MEXICO": "MEX", "SOUTH AFRICA": "RSA", "SOUTH KOREA": "KOR",
    "CZECHIA": "CZE", "CANADA": "CAN", "BOSNIA H.": "BOS",
    "QATAR": "QAT", "SWITZERLAND": "SUI", "BRAZIL": "BRA",
    "MOROCCO": "MAR", "HAITI": "HAI", "SCOTLAND": "SCO",
    "USA": "USA", "PARAGUAY": "PAR", "AUSTRALIA": "AUS",
    "TURKEY": "TUR", "GERMANY": "GER", "CURAÇAO": "CUR",
    "NETHERLANDS": "NED", "JAPAN": "JPN", "SWEDEN": "SWE",
    "TUNISIA": "TUN", "IRAN": "IRN", "NEW ZEALAND": "NZL",
    "BELGIUM": "BEL", "EGYPT": "EGY", "SPAIN": "ESP",
    "CAPO VERDE": "CPV", "SAUDI ARABIA": "KSA", "URUGUAY": "URU",
    "FRANCE": "FRA", "SENEGAL": "SEN", "IRAQ": "IRQ",
    "NORWAY": "NOR", "ARGENTINA": "ARG", "ALGERIA": "ALG",
    "AUSTRIA": "AUT", "JORDAN": "JOR", "IVORY COAST": "CIV",
    "ECUADOR": "ECU", "PORTUGAL": "POR", "DR CONGO": "COD",
    "UZBEKISTAN": "UZB", "COLOMBIA": "COL", "ENGLAND": "ENG",
    "CROATIA": "CRO", "GHANA": "GHA", "PANAMA": "PAN",
}

def display(excel_name):
    """Return the readable display name for a ticker entry."""
    return DISPLAY_NAMES.get(excel_name, excel_name.title())

def short(excel_name):
    """Return 3-letter code for next-game HUD."""
    return SHORT_CODES.get(excel_name, excel_name[:3])


# ─── Fetch helpers ────────────────────────────────────────────────────────────

def _get_json(url, headers=None, timeout=20):
    """Make a GET request, return parsed JSON or raise."""
    req = urllib.request.Request(url, headers={
        "User-Agent": "fantamondiale-bot/1.0",
        **(headers or {})
    })
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read().decode())


def fetch_from_footballdata(api_key):
    """
    Fetch finished group matches from football-data.org.
    Returns (finished_matches, next_game_dict_or_None, is_live).
    """
    print(f"[fetch] football-data.org → {FOOTBALLDATA_URL}")
    try:
        data = _get_json(FOOTBALLDATA_URL, headers={"X-Auth-Token": api_key})
    except urllib.error.HTTPError as e:
        if e.code == 404:
            print(f"  [WARN] football-data.org: WC 2026 not found (404). Key may not have access yet.")
        raise

    finished = []
    upcoming = []
    for m in data.get("matches", []):
        if m.get("stage") not in ("GROUP_STAGE",):
            continue
        status = m.get("status")
        home_api = m.get("homeTeam", {}).get("name", "")
        away_api = m.get("awayTeam", {}).get("name", "")
        home_xl = FOOTBALLDATA_TEAM_MAP.get(home_api)
        away_xl = FOOTBALLDATA_TEAM_MAP.get(away_api)
        if not home_xl or not away_xl:
            if status in ("FINISHED", "LIVE", "IN_PLAY"):
                print(f"  [WARN] Unknown team: '{home_api}' or '{away_api}' — add to FOOTBALLDATA_TEAM_MAP")
            continue
        if status == "FINISHED":
            hs = m.get("score", {}).get("fullTime", {}).get("home")
            as_ = m.get("score", {}).get("fullTime", {}).get("away")
            if hs is None or as_ is None:
                continue
            finished.append({
                "id": str(m.get("id", "")),
                "home": home_xl, "away": away_xl,
                "home_score": int(hs), "away_score": int(as_),
            })
        else:
            is_live = status in ("LIVE", "IN_PLAY", "PAUSED")
            hs, as_ = None, None
            if is_live:
                score_obj = m.get("score", {})
                for time_key in ("fullTime", "regularTime", "halfTime"):
                    hs = score_obj.get(time_key, {}).get("home")
                    as_ = score_obj.get(time_key, {}).get("away")
                    if hs is not None and as_ is not None:
                        break
            upcoming.append({
                "home": home_xl,
                "away": away_xl,
                "utc_date": m.get("utcDate", ""),
                "is_live": is_live,
                "home_score": int(hs) if (is_live and hs is not None) else 0,
                "away_score": int(as_) if (is_live and as_ is not None) else 0,
            })

    print(f"[fetch] Got {len(finished)} finished group matches from football-data.org")
    live_games = []
    next_game = None

    upcoming.sort(key=lambda x: x["utc_date"] or "")
    live_matches = [u for u in upcoming if u["is_live"]]
    future_matches = [u for u in upcoming if not u["is_live"]]

    if live_matches:
        live_games = live_matches[:2]
    if future_matches:
        next_game = future_matches[0]

    if next_game:
        print(f"[fetch] Next game: {short(next_game['home'])}-{short(next_game['away'])}")
    for lg in live_games:
        print(f"[fetch] Live game in progress: {short(lg['home'])}-{short(lg['away'])}")
    return finished, next_game, live_games


def fetch_from_worldcup26():
    """
    Fetch finished group matches from worldcup26.ir (local-only fallback).
    Returns (finished_matches, next_game_dict_or_None, is_live).
    """
    print(f"[fetch] worldcup26.ir → {WORLDCUP26_URL}")
    data = _get_json(WORLDCUP26_URL, timeout=15)
    finished = []
    upcoming = []
    for g in data.get("games", []):
        if g.get("type") != "group":
            continue
        home_api = g.get("home_team_name_en", "")
        away_api = g.get("away_team_name_en", "")
        home_xl = WORLDCUP26_TEAM_MAP.get(home_api)
        away_xl = WORLDCUP26_TEAM_MAP.get(away_api)
        if not home_xl or not away_xl:
            if g.get("finished", "").upper() == "TRUE":
                print(f"  [WARN] Unknown team: '{home_api}' or '{away_api}' — add to WORLDCUP26_TEAM_MAP")
            continue
        if g.get("finished", "").upper() == "TRUE":
            try:
                hs = int(g["home_score"])
                as_ = int(g["away_score"])
            except (KeyError, ValueError, TypeError):
                continue
            finished.append({
                "id": g["id"],
                "home": home_xl, "away": away_xl,
                "home_score": hs, "away_score": as_,
            })
        else:
            hs = g.get("home_score")
            as_ = g.get("away_score")
            is_live = False
            hs_val, as_val = 0, 0
            if hs is not None and as_ is not None and str(hs).strip().lower() not in ("", "null", "none") and str(as_).strip().lower() not in ("", "null", "none"):
                try:
                    hs_val = int(hs)
                    as_val = int(as_)
                    is_live = True
                except ValueError:
                    pass
            upcoming.append({
                "home": home_xl,
                "away": away_xl,
                "date": g.get("local_date") or g.get("date") or g.get("utc_date") or "",
                "is_live": is_live,
                "home_score": hs_val,
                "away_score": as_val,
            })

    print(f"[fetch] Got {len(finished)} finished group matches from worldcup26.ir")
    live_games = []
    next_game = None

    upcoming.sort(key=lambda x: x["date"] or "")
    live_matches = [u for u in upcoming if u["is_live"]]
    future_matches = [u for u in upcoming if not u["is_live"]]

    if live_matches:
        live_games = live_matches[:2]
    if future_matches:
        next_game = future_matches[0]

    if next_game:
        print(f"[fetch] Next game: {short(next_game['home'])}-{short(next_game['away'])}")
    for lg in live_games:
        print(f"[fetch] Live game in progress: {short(lg['home'])}-{short(lg['away'])}")
    return finished, next_game, live_games


def fetch_games():
    """
    Try football-data.org first (works in GitHub Actions).
    Fall back to worldcup26.ir (works locally).
    Returns (finished_list, next_game_or_None, live_games_list).
    """
    from dotenv import load_dotenv
    load_dotenv()

    api_key = os.environ.get("FOOTBALL_DATA_API_KEY", "")
    if api_key:
        try:
            finished, next_game, live_games = fetch_from_footballdata(api_key)
            return finished, next_game, live_games
        except Exception as e:
            print(f"  [WARN] football-data.org failed: {e}")
            print("  [WARN] Falling back to worldcup26.ir ...")

    # Fallback
    try:
        return fetch_from_worldcup26()
    except Exception as e:
        sys.exit(f"[fetch] ERROR: Both APIs failed. Last error: {e}\n"
                 f"  → Set FOOTBALL_DATA_API_KEY env var for reliable cloud access.\n"
                 f"  → Free key: https://www.football-data.org/client/register")


# ─── Excel update ─────────────────────────────────────────────────────────────

def build_excel_match_index(ws):
    from openpyxl.utils import column_index_from_string as ci
    index = {}
    for g, base in GH.items():
        for i in range(1, 7):
            r = base + i
            home = ws.cell(r, ci("B")).value
            away = ws.cell(r, ci("C")).value
            if home and away:
                index[(str(home).strip().upper(), str(away).strip().upper())] = r
    return index


def update_excel(finished_games, dry_run=False):
    import openpyxl
    from openpyxl.utils import column_index_from_string as ci

    if not os.path.exists(MODEL):
        sys.exit(f"[excel] ERROR: Model not found at {MODEL}")

    wb = openpyxl.load_workbook(MODEL)
    ws = wb["Bracket"] if "Bracket" in wb.sheetnames else wb.active
    match_index = build_excel_match_index(ws)

    updated = 0
    skipped = 0
    result_strings = []

    for m in finished_games:
        key = (m["home"], m["away"])
        row = match_index.get(key)
        if row is None:
            print(f"  [WARN] Match not found in Excel: {m['home']} vs {m['away']}")
            skipped += 1
            continue

        current_home = ws.cell(row, ci("D")).value
        current_away = ws.cell(row, ci("E")).value

        result_strings.append(f"{display(m['home']).upper()} {m['home_score']}-{m['away_score']} {display(m['away']).upper()}")

        if current_home == m["home_score"] and current_away == m["away_score"]:
            continue  # already correct

        if dry_run:
            print(f"  [DRY-RUN] Would set row {row}: {m['home']} {m['home_score']}-{m['away_score']} {m['away']}")
        else:
            ws.cell(row, ci("D")).value = m["home_score"]
            ws.cell(row, ci("E")).value = m["away_score"]
            print(f"  [excel] Updated row {row}: {m['home']} {m['home_score']}-{m['away_score']} {m['away']}")
        updated += 1

    if not dry_run and updated > 0:
        wb.save(MODEL)
        print(f"[excel] Saved {MODEL} ({updated} updated, {skipped} skipped).")
    elif dry_run:
        print(f"[dry-run] Would update {updated} match(es), skip {skipped}.")
    else:
        print(f"[excel] No new results to write ({skipped} skipped).")

    return updated, result_strings


# ─── Ticker & Supabase ────────────────────────────────────────────────────────

def build_ticker_text(result_strings):
    if not result_strings:
        return ""
    return "RECENT RESULTS: " + " / ".join(result_strings)

def build_next_game_text(next_game, live_games=None):
    next_game_part = ""
    if next_game:
        next_game_part = f"{short(next_game['home'])}-{short(next_game['away'])} UP NEXT"

    if live_games:
        live_parts = []
        for lg in live_games:
            hs = lg.get("home_score", 0)
            as_ = lg.get("away_score", 0)
            live_parts.append(f"LIVE: {short(lg['home'])} {hs}-{as_} {short(lg['away'])}")

        if len(live_parts) >= 2:
            return f"{live_parts[0]}|{live_parts[1]}"
        elif len(live_parts) == 1:
            if next_game_part:
                return f"{next_game_part}|{live_parts[0]}"
            return f"TBD UP NEXT|{live_parts[0]}"

    return next_game_part if next_game_part else None


def push_meta_to_supabase(ticker_text=None, next_game_text=None):
    """Update ticker_text (and optionally next_game_text) in Supabase."""
    from dotenv import load_dotenv
    load_dotenv()

    url_base = os.environ.get("SUPABASE_URL", "https://ecqieaselexhcqkwbtcy.supabase.co").rstrip("/")
    key = os.environ.get("SUPABASE_SERVICE_KEY")
    if not key:
        print("[supabase] Skipping: SUPABASE_SERVICE_KEY not set.")
        return

    payload = {}
    if ticker_text is not None:
        payload["ticker_text"] = ticker_text
    if next_game_text is not None:
        payload["next_game_text"] = next_game_text

    if not payload:
        print("[supabase] Nothing to update.")
        return

    url = f"{url_base}/rest/v1/xl_leaderboard_meta?id=eq.1"
    body = json.dumps(payload).encode()
    req = urllib.request.Request(url, data=body, method="PATCH", headers={
        "apikey": key,
        "Authorization": "Bearer " + key,
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    })
    try:
        with urllib.request.urlopen(req) as r:
            print(f"[supabase] Updated ticker + next game ✓ (HTTP {r.status})")
    except urllib.error.HTTPError as e:
        print(f"[supabase] ERROR: HTTP {e.code}: {e.read().decode()[:200]}")


def run_subprocess(script_name):
    script = os.path.join(SCRIPT_DIR, script_name)
    print(f"\n[run] python3 {script_name}")
    result = subprocess.run([sys.executable, script], cwd=SCRIPT_DIR)
    if result.returncode != 0:
        print(f"[run] WARNING: {script_name} exited with code {result.returncode}")
    return result.returncode


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Fetch FIFA WC2026 results and update leaderboard.")
    parser.add_argument("--dry-run",      action="store_true", help="Show what would change without writing")
    parser.add_argument("--no-push",      action="store_true", help="Update Excel but skip push_to_supabase")
    parser.add_argument("--ticker-only",  action="store_true", help="Only update ticker text (no Excel write)")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"  FANTAMONDIALE — Auto Result Fetcher")
    print(f"  {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    finished, next_game, live_games = fetch_games()
    print(f"[fetch] {len(finished)} finished group-stage matches found.\n")

    next_game_text = build_next_game_text(next_game, live_games)
    if next_game_text:
        print(f"[next]  {next_game_text}")

    if args.ticker_only:
        ticker = None
        if finished:
            strs = [
                f"{display(m['home']).upper()} {m['home_score']}-{m['away_score']} {display(m['away']).upper()}"
                for m in finished
            ]
            strs.reverse()  # most recent first
            ticker = build_ticker_text(strs[:5])
            print(f"[ticker] {ticker}")
        else:
            print("[ticker] No finished matches found. Skip ticker text generation.")

        if not args.no_push:
            push_meta_to_supabase(ticker, next_game_text)
        else:
            print("[skip] Skip pushing metadata (--no-push)")
        return

    updated = 0
    result_strings = []
    if finished:
        updated, result_strings = update_excel(finished, dry_run=args.dry_run)
    else:
        print("[excel] No finished matches found. Skip Excel update.")

    if args.dry_run:
        print("\n[dry-run] Done. No files written.")
        return

    ticker = None
    if finished and result_strings:
        # Reverse for most-recent first, keep last 5
        res_copy = list(result_strings)
        res_copy.reverse()
        ticker = build_ticker_text(res_copy[:5])
        print(f"\n[ticker] {ticker}")

    if not args.no_push:
        # Update ticker + next game in Supabase
        push_meta_to_supabase(ticker, next_game_text)
        
        # Trigger leaderboard re-computation ONLY if a match actually transitioned to finished in Excel
        if updated > 0:
            print("\n[automation] Match completed and Excel updated. Re-computing leaderboard...")
            run_subprocess("push_to_supabase.py")
        else:
            print("\n[automation] No new games completed (updated = 0). Leaderboard re-computation skipped.")
    else:
        print("\n[skip] Skipping push_to_supabase.py and metadata updates (--no-push)")

    print("\n[done] All updates complete.")


if __name__ == "__main__":
    main()
