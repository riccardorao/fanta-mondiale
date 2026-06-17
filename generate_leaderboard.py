#!/usr/bin/env python3
"""
FIFAWC2026 — automated leaderboard generator.

Reads OFFICIAL results from FIFAWC2026_Model.xlsx (truth) and every participant
file in Pronostici/, scores them, and writes a single self-contained, beautiful
HTML leaderboard (FIFAWC2026_Leaderboard.html) you can share directly.

Run:  python3 generate_leaderboard.py
Re-run any time after updating results in the Model — the HTML regenerates.

SCORING (per the agreed rules):
  GROUP STAGE (per played match, i.e. both goals present in the Model):
    - Correct Score : +5 for EACH team whose exact goals match (perfect = 10)
    - Correct 1X2   : +10 if the result sign matches (independent, stacks)
  GROUP POSITIONS: +10 per correct final-table slot, ONLY once ALL 72
    group-stage matches have been played (gated tournament-wide, not
    per-group) and that group's standings are fully entered (col G).
  KNOCKOUTS / FINAL STANDINGS / TOP SCORER: auto-activate when their truth
    cells are populated in the Model (gated; 0 until then).
"""
import os, glob, json, datetime
from collections import Counter
import openpyxl
from openpyxl.utils import column_index_from_string as ci

# Paths — relative to this script's directory (so you can run from anywhere)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DESKTOP = os.path.expanduser("~/Desktop/FIFAWC2026")
BASE = BASE_DESKTOP if os.path.exists(BASE_DESKTOP) else SCRIPT_DIR
MODEL = os.path.join(BASE, "FIFAWC2026_Model.xlsx")
PRON = os.path.join(BASE, "Pronostici")
OUT = os.path.join(BASE, "FIFAWC2026_Leaderboard.html")

# -------------------------------------------------------------- POINTS (edit) --
POINTS = {
    "score_per_team": 5, "outcome": 10, "position": 10,
    "r32_correct": 10, "r32_wrong": 5,
    "r16_correct": 15, "r16_wrong": 10,
    "qf_correct": 30,  "qf_wrong": 15,
    "sf_correct": 60,  "sf_wrong": 25,
    "final_correct": 90, "final_wrong": 40,
    "standing_1st": 100, "standing_2nd": 60, "standing_3rd": 40, "standing_4th": 20,
    "topscorer_player": 80, "topscorer_goals": 20,
}

# Map the file surname -> full "Name Surname" shown on the leaderboard.
# Fill these in; any surname left out falls back to just the surname.
# (RaoR = Riccardo Rao / RaoT = Tommaso Rao were disambiguated earlier.)
FULL_NAMES = {
    "Antonini": "Stefano Antonini",
    "Bechelli": "Lorenzo Bechelli",
    "Bonaffini": "Edoardo Bonaffini",                 # first name unknown
    "Bose": "Shatanik Bose",
    "Cairoli": "Alessandro Cairoli",
    "Capriotti": "Alessandro Capriotti",
    "Castelli": "Giovanni Castelli",
    "Cernotto": "Giovanni Cernotto",
    "DAmbra": "Francesco D'Ambra",
    "Lima": "Mayara Lima",                           # first name unknown
    "Marchini": "Saverio Marchini",
    "Martinoli": "Francesco Martinoli",
    "Mastrorilli": "Marco Mastrorilli",
    "Parravicini": "Pierpaolo Parravicini",
    "Piergallini": "Eugenio Piergallini",             # first name unknown
    "Polletta": "Matteo Polletta",
    "Priandi": "Bobby Priandi",
    "RaoR": "Riccardo Rao",
    "RaoT": "Tommaso Rao",
    "Redaelli": "Daniele Redaelli",
    "Reuters": "Franziska Reuter",
    "Rigoni": "Alessandro Rigoni",
    "Rola": "Davide Rola",
    "Roncaglia": "Gianluca Roncaglia",
    "Rotellini": "Giorgio Rotellini",                 # first name unknown
    "Silecchia": "Ambrogio Silecchia",
    "Staderini": "Lorenzo Staderini",
    "Stampone": "Valentino Stampone",
    "Trabona": "Tommaso Trabona",
}

# Mapping of manually-written variations to canonical top scorer names for reconciliation.
SCORER_RECONCILIATION_MAP = {
    # Add any variations here if they crop up. E.g.:
    # "KYLIAN MBAPPE": "MBAPPE",
    # "K. MBAPPE": "MBAPPE",
}

def reconcile_scorer_name(name):
    if not name:
        return ""
    # 1. Clean whitespace and convert to uppercase
    name = str(name).strip().upper()
    
    # 2. Check explicit mapping
    if name in SCORER_RECONCILIATION_MAP:
        return SCORER_RECONCILIATION_MAP[name]
        
    # 3. Simple automatic normalization: remove accents/diacritics
    import unicodedata
    name = "".join(
        c for c in unicodedata.normalize("NFD", name)
        if unicodedata.category(c) != "Mn"
    )
    
    # Remove common prefixes like "K. ", "C. ", "L. " (initials)
    import re
    name = re.sub(r'^[A-Z]\.\s+', '', name)
    
    # Re-check mapping after normalization
    if name in SCORER_RECONCILIATION_MAP:
        return SCORER_RECONCILIATION_MAP[name]
        
    return name

GH = {g: r for g, r in zip("ABCDEFGHIJKL", [4, 11, 18, 25, 32, 39, 46, 53, 60, 67, 74, 81])}

# ---- Tournament structure (2026: 48 teams, 12 groups of 4) ----
N_GROUPS = 12
GROUP_MATCHES_TOTAL = 72          # 12 groups x C(4,2)=6
KO_SLOTS = {"r32": 32, "r16": 16, "qf": 8, "sf": 4, "final": 2}  # teams reaching each round

KO_ROUNDS = {
    "r32": {"col": 11, "rows": [5, 6, 9, 10, 13, 14, 17, 18, 21, 22, 25, 26, 29, 30, 33, 34, 37, 38, 41, 42, 45, 46, 49, 50, 53, 54, 57, 58, 61, 62, 65, 66]},
    "r16": {"col": 16, "rows": [7, 8, 15, 16, 23, 24, 31, 32, 39, 40, 47, 48, 55, 56, 63, 64]},
    "qf": {"col": 21, "rows": [11, 12, 27, 28, 43, 44, 59, 60]},
    "sf": {"col": 26, "rows": [19, 20, 51, 52]},
    "final": {"col": 31, "rows": [35, 36]},
}

STANDINGS_MAP = {
    35: "standing_1st",
    36: "standing_2nd",
    37: "standing_3rd",
    38: "standing_4th",
}


def tournament_max(P):
    """Maximum points achievable across the WHOLE tournament (best case)."""
    group = (GROUP_MATCHES_TOTAL * 2 * P["score_per_team"]      # 720  per-team exact
             + GROUP_MATCHES_TOTAL * P["outcome"]               # 720  outcome
             + N_GROUPS * 4 * P["position"])                    # 480  positions
    ko = (KO_SLOTS["r32"] * P["r32_correct"] + KO_SLOTS["r16"] * P["r16_correct"]
          + KO_SLOTS["qf"] * P["qf_correct"] + KO_SLOTS["sf"] * P["sf_correct"]
          + KO_SLOTS["final"] * P["final_correct"])             # 1220
    standings = P["standing_1st"] + P["standing_2nd"] + P["standing_3rd"] + P["standing_4th"]  # 220
    topscorer = P["topscorer_player"] + P["topscorer_goals"]    # 100
    return group + ko + standings + topscorer                  # 3460 with defaults


def norm(v):
    return str(v).strip().upper() if v not in (None, "") else None


def outcome(h, a):
    if h is None or a is None:
        return None
    try:
        h = float(h); a = float(a)
    except (TypeError, ValueError):
        return None
    return "H" if h > a else ("A" if h < a else "D")


def is_num(v):
    if v is None or v == "":
        return False
    try:
        float(v); return True
    except (TypeError, ValueError):
        return False


def read_truth(ws):
    """Returns truth: matches, group standings, knockout rounds truth, standings truth, top scorer truth."""
    matches = []       # (row, gh, ga, outcome)
    raw_positions = {} # group -> {row: team}  (only if all 4 slots present)
    for g, base in GH.items():
        for i in range(1, 7):
            r = base + i
            d, e = ws.cell(r, ci("D")).value, ws.cell(r, ci("E")).value
            if is_num(d) and is_num(e):
                matches.append((r, float(d), float(e), outcome(d, e)))
        slots = {}
        for i in range(1, 5):
            r = base + i
            t = norm(ws.cell(r, ci("G")).value)
            if t:
                slots[r] = t
        if len(slots) == 4:          # group standings finalised
            raw_positions[g] = slots
    # Group positions only count once the ENTIRE group stage is played —
    # partial standings (predicted, not yet locked in) don't score early.
    group_stage_complete = len(matches) == GROUP_MATCHES_TOTAL
    positions = raw_positions if group_stage_complete else {}

    ko_truth = {}
    standings_truth = {}
    topscorer_player_truth = None
    topscorer_goals_truth = None

    if group_stage_complete:
        # Extract knockout truth (only non-empty cells count)
        for round_name, info in KO_ROUNDS.items():
            col = info["col"]
            ko_truth[round_name] = {}
            for r in info["rows"]:
                val = norm(ws.cell(r, col).value)
                if val:
                    ko_truth[round_name][r] = val

        # Extract standings truth
        for r in STANDINGS_MAP.keys():
            val = norm(ws.cell(r, ci("AJ")).value)
            if val:
                standings_truth[r] = val

        # Extract top scorer truth
        topscorer_player_truth = norm(ws.cell(44, ci("AJ")).value)
        topscorer_goals_truth = ws.cell(44, ci("AK")).value
        if topscorer_goals_truth is not None and not is_num(topscorer_goals_truth):
            topscorer_goals_truth = None

    return matches, positions, ko_truth, standings_truth, topscorer_player_truth, topscorer_goals_truth


def score_file(path, matches, positions, ko_truth=None, standings_truth=None, topscorer_player_truth=None, topscorer_goals_truth=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Bracket"] if "Bracket" in wb.sheetnames else wb.active
    bd = {"Correct Score": 0, "Correct Outcome": 0, "Group Positions": 0,
          "Knockouts": 0, "Final Standings": 0, "Top Scorer": 0}
    # group scores + outcome
    for (r, gh, ga, gout) in matches:
        ph, pa = ws.cell(r, ci("D")).value, ws.cell(r, ci("E")).value
        if not (is_num(ph) and is_num(pa)):
            continue
        ph, pa = float(ph), float(pa)
        if ph == gh:
            bd["Correct Score"] += POINTS["score_per_team"]
        if pa == ga:
            bd["Correct Score"] += POINTS["score_per_team"]
        if outcome(ph, pa) == gout:
            bd["Correct Outcome"] += POINTS["outcome"]
    # group positions (only completed groups)
    for g, slots in positions.items():
        for r, true_team in slots.items():
            if norm(ws.cell(r, ci("G")).value) == true_team:
                bd["Group Positions"] += POINTS["position"]

    # knockouts (gated by populated model truth)
    if ko_truth:
        for round_name, truth_slots in ko_truth.items():
            col = KO_ROUNDS[round_name]["col"]
            correct_key = f"{round_name}_correct"
            wrong_key = f"{round_name}_wrong"
            for r, true_team in truth_slots.items():
                pred_team = norm(ws.cell(r, col).value)
                if pred_team == true_team:
                    bd["Knockouts"] += POINTS[correct_key]
                else:
                    bd["Knockouts"] += POINTS[wrong_key]

    # final standings
    if standings_truth:
        for r, true_team in standings_truth.items():
            pred_team = norm(ws.cell(r, ci("AJ")).value)
            if pred_team == true_team:
                key = STANDINGS_MAP[r]
                bd["Final Standings"] += POINTS[key]

    # top scorer
    if topscorer_player_truth:
        pred_player = norm(ws.cell(44, ci("AI")).value)
        if reconcile_scorer_name(pred_player) == reconcile_scorer_name(topscorer_player_truth):
            bd["Top Scorer"] += POINTS["topscorer_player"]

    if topscorer_goals_truth is not None:
        pred_goals = ws.cell(44, ci("AK")).value
        if pred_goals is not None:
            try:
                if float(pred_goals) == float(topscorer_goals_truth):
                    bd["Top Scorer"] += POINTS["topscorer_goals"]
            except (TypeError, ValueError):
                pass

    # predicted winner from cell AJ35
    predicted_winner = norm(ws.cell(35, ci("AJ")).value)
    if predicted_winner == "WINNER":
        predicted_winner = None

    # predicted top scorer from cell AI44
    predicted_scorer = ws.cell(44, ci("AI")).value
    if predicted_scorer:
        s_norm = norm(predicted_scorer)
        if s_norm == "PLAYER NAME":
            predicted_scorer = None
        else:
            reconciled = reconcile_scorer_name(predicted_scorer)
            predicted_scorer = " ".join([w.capitalize() for w in reconciled.split()]) if reconciled else None
    else:
        predicted_scorer = None

    total = sum(bd.values())
    return total, bd, predicted_winner, predicted_scorer


def compute_leaderboard_data(model_path, pron_dir):
    """Loads model, extracts truth, scores files, and computes ranks + metadata."""
    wbM = openpyxl.load_workbook(model_path, data_only=True)
    wsM = wbM["Bracket"] if "Bracket" in wbM.sheetnames else wbM.active
    matches, positions, ko_truth, standings_truth, topscorer_player_truth, topscorer_goals_truth = read_truth(wsM)

    rows = []
    predicted_winners = []
    predicted_scorers = []
    for f in sorted(glob.glob(os.path.join(pron_dir, "FIFAWC2026_*.xlsx"))):
        name = os.path.basename(f).replace("FIFAWC2026_", "").replace(".xlsx", "")
        display = FULL_NAMES.get(name, name)
        try:
            total, bd, pred_w, pred_s = score_file(
                f, matches, positions, ko_truth, standings_truth, 
                topscorer_player_truth, topscorer_goals_truth
            )
            rows.append({
                "name": display, "key": name, "total": total, "bd": bd, 
                "predicted_winner": pred_w
            })
            if pred_w:
                predicted_winners.append(pred_w)
            if pred_s:
                predicted_scorers.append(pred_s)
        except Exception as ex:
            rows.append({
                "name": display, "key": name, "total": -1, "bd": {}, 
                "error": str(ex), "predicted_winner": None
            })

    winner_counts = Counter(predicted_winners).most_common()
    scorer_counts = Counter(predicted_scorers).most_common(5)

    stats_winners = [{"team": w, "count": count} for w, count in winner_counts]
    stats_scorers = [{"player": s, "count": count} for s, count in scorer_counts]

    # ---- maximum points ANYONE could have earned up to the current round ----
    max_possible = len(matches) * (2 * POINTS["score_per_team"] + POINTS["outcome"])
    max_possible += len(positions) * 4 * POINTS["position"]
    for round_name, truth_slots in ko_truth.items():
        max_possible += len(truth_slots) * POINTS[f"{round_name}_correct"]
    for r in standings_truth.keys():
        key = STANDINGS_MAP[r]
        max_possible += POINTS[key]
    if topscorer_player_truth:
        max_possible += POINTS["topscorer_player"]
    if topscorer_goals_truth is not None:
        max_possible += POINTS["topscorer_goals"]

    max_possible = max(max_possible, 1)

    rows.sort(key=lambda x: -x["total"])
    # dense ranking
    rank = 0; prev = None
    for i, r in enumerate(rows):
        if r["total"] != prev:
            rank = i + 1; prev = r["total"]
        r["rank"] = rank

    meta = {
        "generated": datetime.datetime.now().strftime("%d %b %Y, %H:%M"),
        "matches_played": len(matches),
        "groups_complete": len(positions),
        "participants": len(rows),
        "max_total": max((r["total"] for r in rows if r["total"] >= 0), default=0) or 1,
        "max_possible": max_possible,
        "tournament_max": tournament_max(POINTS),
        "stats": {
            "winners": stats_winners,
            "scorers": stats_scorers
        }
    }
    return rows, meta


def main():
    rows, meta = compute_leaderboard_data(MODEL, PRON)
    html = render_html(rows, meta)
    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write(html)
    print("Wrote", OUT)
    print(f"{meta['participants']} participants, {meta['matches_played']} matches scored")
    for r in rows[:5]:
        print(f"  #{r['rank']:<2} {r['name']:<14} {r['total']}")


# ----------------------------------------------------------------- HTML render -
def render_html(rows, meta):
    data_json = json.dumps(rows)
    meta_json = json.dumps(meta)
    cats = ["group scores", "group outcomes", "group standings",
            "knockout phase", "final standings", "top scorer"]
    cats_json = json.dumps(cats)
    return TEMPLATE.replace("__DATA__", data_json)\
                   .replace("__META__", meta_json)\
                   .replace("__CATS__", cats_json)


TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>FIFA WC 2026 · Fantamondiale Leaderboard</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Press+Start+2P&family=VT323&display=swap');
  :root{
    --neon-green:#39FF14; --neon-pink:#FF10F0; --neon-cyan:#00FFFF; --neon-yellow:#FFE800;
    --gold:#FFD700; --silver:#C0C0C0; --bronze:#CD7F32;
    --ink:#050014;
  }
  html{background:#000}
  *{box-sizing:border-box;margin:0;padding:0;border-radius:0!important}
  body{
    font-family:'VT323',monospace;font-size:18px;
    background:
      repeating-linear-gradient(0deg, rgba(57,255,20,0.05) 0px, rgba(57,255,20,0.05) 1px, transparent 1px, transparent 40px),
      repeating-linear-gradient(90deg, rgba(57,255,20,0.05) 0px, rgba(57,255,20,0.05) 1px, transparent 1px, transparent 40px),
      radial-gradient(circle at 50% 0%, #1a0033 0%, #0a0020 55%, #000 100%);
    color:var(--neon-green);min-height:100vh;padding:18px 14px 60px;
    position:relative;overflow-x:hidden;
  }
  /* CRT scanline + flicker overlay */
  body::before{
    content:'';position:fixed;inset:0;pointer-events:none;z-index:9999;
    background:repeating-linear-gradient(0deg,rgba(0,0,0,.25) 0px,rgba(0,0,0,.25) 1px,transparent 1px,transparent 3px);
    mix-blend-mode:overlay;
  }
  body::after{
    content:'';position:fixed;inset:0;pointer-events:none;z-index:9998;
    background:radial-gradient(ellipse at center,transparent 55%,rgba(0,0,0,.55) 100%);
    animation:flicker 6s infinite;
  }
  @keyframes flicker{0%,100%{opacity:1}92%{opacity:1}93%{opacity:.82}94%{opacity:1}97%{opacity:1}98%{opacity:.88}}
  .wrap{max-width:1080px;margin:0 auto;position:relative}

  /* HUD bar */
  .hud{display:flex;justify-content:space-between;flex-wrap:wrap;gap:10px;
    font-family:'Press Start 2P',monospace;font-size:10px;margin-bottom:22px;
    border:2px solid var(--neon-cyan);box-shadow:0 0 0 3px var(--ink),0 0 0 5px var(--neon-pink),0 0 16px rgba(0,255,255,.4);
    padding:10px 16px;background:rgba(5,0,20,0.6)}
  .hud-item{color:var(--neon-yellow);text-shadow:0 0 6px var(--neon-yellow)}
  .hud-item.blink{animation:blink 1s steps(2) infinite}
  @keyframes blink{50%{opacity:.25}}

  header{text-align:center;margin-bottom:8px}
  h1{font-family:'Press Start 2P',monospace;font-size:clamp(18px,5.5vw,50px);line-height:1.2;
    color:var(--neon-yellow);letter-spacing:3px;font-weight:900;
    text-shadow:3px 3px 0 var(--neon-pink),6px 6px 0 var(--neon-cyan),0 0 24px rgba(255,232,0,.5);
    animation:glitchIn .6s ease-out}
  @keyframes glitchIn{
    0%{opacity:0;transform:translate(-12px,0) skewX(8deg);filter:hue-rotate(120deg)}
    25%{opacity:1;transform:translate(8px,0) skewX(-6deg)}
    50%{transform:translate(-4px,0) skewX(3deg);filter:hue-rotate(40deg)}
    75%{transform:translate(2px,0) skewX(-1deg);filter:hue-rotate(0deg)}
    100%{transform:translate(0,0) skewX(0);opacity:1}
  }
  .ticker{margin:14px auto 0;max-width:680px;overflow:hidden;white-space:nowrap;
    border-top:2px solid var(--neon-pink);border-bottom:2px solid var(--neon-pink);padding:6px 0}
  .ticker span{display:inline-block;padding-left:100%;font-size:15px;color:var(--neon-cyan);
    text-shadow:0 0 6px var(--neon-cyan);animation:marquee 16s linear infinite}
  @keyframes marquee{to{transform:translateX(-100%)}}
  .sub{display:none}
  .stats{display:none}

  .section-label{text-align:center;font-family:'Press Start 2P',monospace;font-size:13px;
    color:var(--neon-pink);text-shadow:0 0 8px var(--neon-pink);margin:30px 0 16px;letter-spacing:2px}

  .podium{display:flex;gap:18px;justify-content:center;align-items:flex-end;margin:0 0 30px;flex-wrap:wrap}
  .pod{flex:1;min-width:170px;max-width:220px;padding:20px 14px;text-align:center;
    border:2px solid var(--neon-cyan);box-shadow:0 0 0 3px var(--ink),0 0 0 5px var(--neon-pink),0 0 18px rgba(0,255,255,.45);
    background:rgba(5,0,20,0.65);
    transform:translateY(20px);opacity:0;animation:rise .6s forwards}
  .pod .medal{font-size:38px;margin-bottom:6px}
  .pod .pname{font-family:'Press Start 2P',monospace;font-size:clamp(12px, 3.2vw, 16px);color:var(--neon-cyan);text-transform:uppercase;letter-spacing:1px;margin:6px 0;line-height:1.3}
  .pod .ppts{font-family:'Press Start 2P',monospace;font-size:26px;letter-spacing:1px;font-weight:900;margin:8px 0;color:var(--neon-green)}
  .pod .plabel{display:none}
  .pod-rank{font-family:'Press Start 2P',monospace;font-size:13px;margin-top:4px;letter-spacing:1px}
  @keyframes rise{to{transform:translateY(0);opacity:1}}
  @keyframes flash{0%,100%{opacity:1;filter:brightness(1)}50%{opacity:.5;filter:brightness(1.8)}}
  .r1{color:var(--gold);text-shadow:0 0 10px var(--gold);animation:flash 1.1s infinite}
  .r2{color:var(--silver);text-shadow:0 0 10px var(--silver);animation:flash 1.3s infinite}
  .r3{color:var(--bronze);text-shadow:0 0 10px var(--bronze);animation:flash 1.5s infinite}

  .board{display:flex;flex-direction:column;gap:10px;margin-top:8px}
  .row{background:rgba(5,0,20,0.6);overflow:hidden;
    border:2px solid var(--neon-cyan);box-shadow:0 0 0 3px var(--ink),0 0 0 5px var(--neon-pink),0 0 14px rgba(0,255,255,.3);
    opacity:0;transform:translateY(8px);animation:fade .5s forwards;transition:box-shadow .2s,background .2s}
  .row:hover{background:rgba(20,0,40,0.85);box-shadow:0 0 0 3px var(--ink),0 0 0 5px var(--neon-green),0 0 22px rgba(57,255,20,.55)}
  @keyframes fade{to{opacity:1;transform:none}}
  .rmain{display:grid;grid-template-columns:22px 52px 1fr auto 32px;align-items:center;gap:10px;
    padding:13px 16px;cursor:pointer;transition:transform .1s}
  .rmain:active{transform:translateY(2px)}
  .ball{font-size:14px;text-align:center;opacity:0;transition:opacity .2s}
  .row:hover .ball{opacity:1;animation:bob .5s ease-in-out infinite alternate}
  @keyframes bob{from{transform:translateY(0) rotate(0deg)}to{transform:translateY(-3px) rotate(20deg)}}
  .rank{font-family:'Press Start 2P',monospace;font-size:12px;color:var(--neon-green);text-align:center;text-shadow:0 0 6px var(--neon-green)}
  .who{font-family:'Press Start 2P',monospace;font-size:10px;color:var(--neon-cyan);text-transform:uppercase;letter-spacing:1px}
  .barwrap{height:12px;background:#000;border:1px solid var(--neon-cyan);margin-top:7px;overflow:hidden;width:min(440px,44vw)}
  .bar{height:100%;width:0;transition:width 1.2s cubic-bezier(.2,.8,.2,1);box-shadow:0 0 10px currentColor}
  .pts{font-family:'Press Start 2P',monospace;font-size:14px;font-weight:900;text-align:right;color:var(--neon-yellow)}
  .chev{color:var(--neon-pink);transition:transform .25s;text-align:center;font-weight:900;font-size:12px;text-shadow:0 0 6px var(--neon-pink)}
  .row.open .chev{transform:rotate(180deg)}
  .detail{max-height:0;overflow:hidden;transition:max-height .35s ease}
  .row.open .detail{max-height:420px;border-top:2px solid var(--neon-cyan)}
  .cats{display:grid;grid-template-columns:repeat(2,1fr);gap:10px;padding:16px 16px 4px;background:rgba(0,0,0,0.4)}
  @media (min-width:700px){.cats{grid-template-columns:repeat(6,1fr)}}
  .cat{background:rgba(5,0,20,0.7);border:1px solid var(--neon-cyan);box-shadow:0 0 0 2px var(--ink),0 0 0 3px var(--neon-pink);padding:10px;text-align:center}
  .cat .cl{font-size:14px;color:#fff;font-weight:400;line-height:1.2}
  .cat .cv{font-size:18px;margin-top:6px;color:var(--neon-green);font-weight:700;text-shadow:none}
  .cat.zero{opacity:.35}
  .ctx{display:flex;justify-content:space-between;align-items:center;padding:4px 16px 14px;font-size:14px;color:var(--neon-cyan);background:rgba(0,0,0,0.4)}
  .ctx-winner{text-align:right}
  footer{text-align:center;color:var(--neon-cyan);font-size:13px;margin-top:36px;line-height:1.6;text-shadow:0 0 6px var(--neon-cyan)}
  @media (max-width:699px){
    .podium{flex-direction:column;align-items:stretch;gap:14px}
    .pod{max-width:100%}
    .ball{display:none}
    .rmain{grid-template-columns:40px 1fr auto 28px}
  }
  /* Tabs */
  .tabs{display:flex;justify-content:center;gap:20px;margin:20px 0}
  .tab-btn{font-family:'Press Start 2P',monospace;font-size:12px;background:rgba(5,0,20,0.7);color:var(--neon-cyan);border:2px solid var(--neon-cyan);padding:10px 20px;cursor:pointer;transition:all .2s;text-shadow:0 0 5px var(--neon-cyan);box-shadow:0 0 0 2px var(--ink),0 0 0 4px var(--neon-pink)}
  .tab-btn:hover{color:var(--neon-green);border-color:var(--neon-green);text-shadow:0 0 5px var(--neon-green);box-shadow:0 0 0 2px var(--ink),0 0 0 4px var(--neon-green)}
  .tab-btn.active{color:var(--neon-yellow);border-color:var(--neon-yellow);text-shadow:0 0 8px var(--neon-yellow);box-shadow:0 0 0 2px var(--ink),0 0 0 4px var(--neon-yellow),0 0 15px rgba(255,232,0,.4)}

  /* Stats Panel */
  .stats-container{display:flex;flex-direction:column;gap:30px;margin-top:20px}
  .stats-card{background:rgba(5,0,20,0.65);border:2px solid var(--neon-cyan);box-shadow:0 0 0 3px var(--ink),0 0 0 5px var(--neon-pink),0 0 16px rgba(0,255,255,0.35);padding:24px 20px}
  .stats-title{font-family:'Press Start 2P',monospace;font-size:14px;color:var(--neon-yellow);text-shadow:0 0 8px var(--neon-yellow);text-align:center;margin-bottom:24px;letter-spacing:1px}
  .stats-grid-winners{display:flex;justify-content:center;gap:16px;flex-wrap:wrap}
  .winner-box{flex:1;min-width:150px;max-width:220px;border:1px solid var(--neon-cyan);background:rgba(0,0,0,0.4);padding:16px 10px;text-align:center}
  .winner-box .w-rank{font-family:'Press Start 2P',monospace;font-size:11px;color:var(--neon-pink);margin-bottom:8px}
  .winner-box .w-flag{font-size:44px;margin-bottom:8px}
  .winner-box .w-team{font-family:'Press Start 2P',monospace;font-size:11px;color:var(--neon-cyan);text-transform:uppercase;margin-bottom:6px}
  .winner-box .w-count{font-size:18px;color:var(--neon-green);font-weight:bold}
  .stats-scorers-list{display:flex;flex-direction:column;gap:12px;max-width:600px;margin:0 auto}
  .scorer-row{display:grid;grid-template-columns:30px 1fr auto;align-items:center;gap:12px;background:rgba(0,0,0,0.3);border:1px solid var(--neon-cyan);padding:10px 16px}
  .scorer-rank{font-family:'Press Start 2P',monospace;font-size:11px;color:var(--neon-pink)}
  .scorer-info{display:flex;flex-direction:column}
  .scorer-name{font-family:'Press Start 2P',monospace;font-size:10px;color:var(--neon-cyan);text-transform:uppercase}
  .scorer-bar-wrap{height:8px;background:#000;border:1px solid var(--neon-cyan);margin-top:6px;width:100%}
  .scorer-bar{height:100%;background:var(--neon-green);box-shadow:0 0 8px var(--neon-green);width:0;transition:width 1s ease-out}
  .scorer-votes{font-family:'Press Start 2P',monospace;font-size:11px;color:var(--neon-yellow);text-align:right}
</style>
</head>
<body>
<div class="wrap">
  <div class="hud">
    <span class="hud-item">CREDIT 01</span>
    <span class="hud-item blink">1P READY</span>
    <span class="hud-item">STAGE 1</span>
  </div>
  <header>
    <h1>LEADERBOARD</h1>
    <div class="ticker"><span>★ WELCOME TO THE FANTAMONDIALE ARENA ★ INSERT COIN TO CONTINUE ★ MAY THE BEST PREDICTOR WIN ★ GOOD LUCK PLAYER ★</span></div>
    <div class="sub" id="sub"></div>
    <div class="stats" id="stats"></div>
  </header>
  <div class="tabs">
    <button id="tab-leaderboard" class="tab-btn active" onclick="switchTab('leaderboard')">LEADERBOARD</button>
    <button id="tab-stats" class="tab-btn" onclick="switchTab('stats')">STATISTICS</button>
  </div>
  <div id="view-leaderboard">
    <div class="section-label">★ HIGH SCORE ★</div>
    <div class="podium" id="podium"></div>
    <div class="board" id="board"></div>
  </div>
  <div id="view-stats" style="display: none;">
    <div class="section-label">★ PREDICTION STATISTICS ★</div>
    <div class="stats-container" id="stats-container"></div>
  </div>

  <footer id="foot"></footer>
</div>

<script>
const DATA = __DATA__;
const META = __META__;
const CATS = __CATS__;
const CAT_KEYS = ["Correct Score", "Correct Outcome", "Group Positions",
                  "Knockouts", "Final Standings", "Top Scorer"];
const maxT = Math.max(1, META.max_possible || 1);
const RANK_LABEL = ['1ST','2ND','3RD'];

const COUNTRY_FLAGS = {
  'Algeria': '🇩🇿',
  'Argentina': '🇦🇷',
  'Australia': '🇦🇺',
  'Austria': '🇦🇹',
  'Belgium': '🇧🇪',
  'Bosnia H.': '🇧🇦',
  'Brazil': '🇧🇷',
  'Canada': '🇨🇦',
  'Capo Verde': '🇨🇻',
  'Colombia': '🇨🇴',
  'Congo DR': '🇨🇩',
  'Costa Rica': '🇨🇷',
  'Croatia': '🇭🇷',
  'Curaçao': '🇨🇼',
  'Czechia': '🇨🇿',
  'Ecuador': '🇪🇨',
  'Egypt': '🇪🇬',
  'England': '🏴󠁧󠁢󠁥󠁮󠁧󠁿',
  'France': '🇫🇷',
  'Germany': '🇩🇪',
  'Ghana': '🇬🇭',
  'Haiti': '🇭🇹',
  'Iran': '🇮🇷',
  'Iraq': '🇮🇶',
  'Ivory Coast': '🇨🇮',
  'Japan': '🇯🇵',
  'Jordan': '🇯🇴',
  'Mexico': '🇲🇽',
  'Morocco': '🇲🇦',
  'Netherlands': '🇳🇱',
  'New Zealand': '🇳🇿',
  'Norway': '🇳🇴',
  'Panama': '🇵🇦',
  'Paraguay': '🇵🇾',
  'Portugal': '🇵🇹',
  'Qatar': '🇶🇦',
  'Saudi Arabia': '🇸🇦',
  'Scotland': '🏴󠁧󠁢󠁳󠁣󠁴󠁿',
  'Senegal': '🇸🇳',
  'South Africa': '🇿🇦',
  'South Korea': '🇰🇷',
  'Spain': '🇪🇸',
  'Sweden': '🇸🇪',
  'Switzerland': '🇨🇭',
  'Tunisia': '🇹🇳',
  'Turkey': '🇹🇷',
  'Uruguay': '🇺🇾',
  'USA': '🇺🇸',
  'Uzbekistan': '🇺🇿'
};

function getFlagEmoji(country) {
  if (!country) return '🌍';
  const match = Object.keys(COUNTRY_FLAGS).find(k => k.toLowerCase() === country.toLowerCase());
  return match ? COUNTRY_FLAGS[match] : '🌍';
}

function switchTab(tab) {
  const btnLeaderboard = document.getElementById('tab-leaderboard');
  const btnStats = document.getElementById('tab-stats');
  const viewLeaderboard = document.getElementById('view-leaderboard');
  const viewStats = document.getElementById('view-stats');
  
  if (tab === 'leaderboard') {
    btnLeaderboard.classList.add('active');
    btnLeaderboard.innerText = 'LEADERBOARD';
    btnStats.classList.remove('active');
    btnStats.innerText = 'STATISTICS';
    viewLeaderboard.style.display = 'block';
    viewStats.style.display = 'none';
  } else {
    btnLeaderboard.classList.remove('active');
    btnLeaderboard.innerText = 'LEADERBOARD';
    btnStats.classList.add('active');
    btnStats.innerText = 'STATISTICS';
    viewLeaderboard.style.display = 'none';
    viewStats.style.display = 'block';
    
    setTimeout(() => {
      document.querySelectorAll('.scorer-bar').forEach(b => b.style.width = b.dataset.w + '%');
    }, 50);
  }
}

function renderStats(meta, totalParticipants) {
  const container = document.getElementById('stats-container');
  if (!meta.stats || !meta.stats.winners || !meta.stats.scorers) {
    container.innerHTML = `<div class="state">NO STATISTICS UPLOADED YET</div>`;
    return;
  }
  
  const winners = meta.stats.winners || [];
  const scorers = meta.stats.scorers || [];
  
  // Render Winners (all predicted, with bars relative to the top country)
  let winnersHtml = '';
  if (winners.length > 0) {
    const refCount = winners[0].count || 1;
    winnersHtml = `
      <div class="stats-card">
        <div class="stats-title">MOST PREDICTED ⭐️ WORLD CHAMPIONS ⭐️</div>
        <div class="stats-scorers-list">
    `;
    
    winners.forEach((w, i) => {
      const pct = Math.round((w.count / refCount) * 100);
      const flag = getFlagEmoji(w.team);
      const barColor = `hsl(${120 * (w.count / refCount)}, 100%, 50%)`;
      winnersHtml += `
        <div class="scorer-row">
          <div class="scorer-rank">#${i+1}</div>
          <div class="scorer-info">
            <span class="scorer-name">${w.team} ${flag}</span>
            <div class="scorer-bar-wrap">
              <div class="scorer-bar" data-w="${pct}" style="background:${barColor}; box-shadow: 0 0 10px ${barColor}"></div>
            </div>
          </div>
          <div class="scorer-votes">${w.count} ${w.count === 1 ? 'VOTE' : 'VOTES'}</div>
        </div>
      `;
    });
    
    winnersHtml += `
        </div>
      </div>
    `;
  }
  
  // Render Scorers (top 5, vertical list, no bars, vote count written)
  let scorersHtml = '';
  if (scorers.length > 0) {
    scorersHtml = `
      <div class="stats-card">
        <div class="stats-title">MOST PREDICTED ⚽️ TOP SCORERS ⚽️</div>
        <div class="stats-scorers-list">
    `;
    
    scorers.slice(0, 5).forEach((s, i) => {
      scorersHtml += `
        <div class="scorer-row no-bar">
          <div class="scorer-rank">#${i+1}</div>
          <div class="scorer-info">
            <span class="scorer-name">${s.player}</span>
          </div>
          <div class="scorer-votes">${s.count} ${s.count === 1 ? 'VOTE' : 'VOTES'}</div>
        </div>
      `;
    });
    
    scorersHtml += `
        </div>
      </div>
    `;
  }
  
  container.innerHTML = winnersHtml + scorersHtml;
}

// footer
document.getElementById('foot').innerHTML =
  `© FANTAMONDIALE 2026 - ALL RIGHTS RESERVED<br>UPDATED ${META.generated}`;

// podium (top 3)
const podEl = document.getElementById('podium');
const medal = ['🥇','🥈','🥉'];
DATA.slice(0,3).forEach((d,i)=>{
  const el=document.createElement('div');
  el.className='pod';
  el.innerHTML=`<div class="medal">${medal[i]}</div>
    <div class="pname">${d.name}</div>
    <div class="ppts">${d.total}</div>
    <div class="pod-rank r${i+1}">${RANK_LABEL[i]}</div>`;
  podEl.appendChild(el);
});

// full board
const board=document.getElementById('board');
function build(list){
  board.innerHTML='';
  const maxScore = Math.max(...list.map(x => x.total || 0), 1);
  list.forEach((d,idx)=>{
    const row=document.createElement('div');
    row.className='row'; row.style.animationDelay=(idx*0.02)+'s';
    const catsHtml = CATS.map((c,i)=>{
      const v=(d.bd&&d.bd[CAT_KEYS[i]])||0;
      return `<div class="cat ${v===0?'zero':''}"><div class="cl">${c}</div>
              <div class="cv">${v}</div></div>`;
    }).join('');
    const pctScore = d.total / maxScore;
    const barColor = `hsl(${120 * pctScore}, 100%, 50%)`;
    const pctN = Math.round(100*d.total/maxT);
    const rankHtml = d.rank<=3
      ? `<span class="r${d.rank}">${RANK_LABEL[d.rank-1]}</span>`
      : d.rank;
    row.innerHTML=`
      <div class="rmain">
        <div class="ball">⚽</div>
        <div class="rank">${rankHtml}</div>
        <div><div class="who">${d.name}</div>
          <div class="barwrap"><div class="bar" style="background:${barColor}" data-w="${pctN}"></div></div></div>
        <div class="pts">${d.total}</div>
        <div class="chev">▾</div>
      </div>
      <div class="detail">
        <div class="cats">${catsHtml}</div>
        <div class="ctx">
          <div>${d.total}/${META.max_possible||0} PTS · ${pctN}%</div>
          <div class="ctx-winner">${d.predicted_winner ? 'WORLD CHAMPIONS: ' + d.predicted_winner + ' ' + getFlagEmoji(d.predicted_winner) : ''}</div>
        </div>
      </div>`;
    row.querySelector('.rmain').onclick=()=>row.classList.toggle('open');
    board.appendChild(row);
  });
  requestAnimationFrame(()=>{
    document.querySelectorAll('.bar').forEach(b=>b.style.width=b.dataset.w+'%');
  });
}
build(DATA);
renderStats(META, DATA.length);
</script>
</body>
</html>"""


if __name__ == "__main__":
    main()
